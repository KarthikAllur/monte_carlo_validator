"""
excel_handler.py — Read input queries and write styled output results using pandas/openpyxl.

Designed so this module can be replaced with a SharePoint handler in the
future without changing sql_parser.py or snowflake_connector.py.

Interface contract
──────────────────
    read_queries(filepath: str) -> list[str]
    write_results(filepath: str, results: list[dict]) -> None

A future SharePointHandler module only needs to satisfy the same contract.
"""

import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Column names (single source of truth) ─────────────────────────────────────

INPUT_QUERY_COLUMN: str = "query"

OUTPUT_COLUMNS: list[str] = [
    "query",
    "total_count_query",
    "error_count_query",
    "total_count",
    "error_count",
    "status",
    "error_message",
]

# Friendly display headers shown in Excel
COLUMN_HEADERS: dict[str, str] = {
    "query":             "Monte Carlo Query",
    "total_count_query": "Total Count Query",
    "error_count_query": "Error Count Query",
    "total_count":       "Total Count",
    "error_count":       "Error Count",
    "status":            "Status",
    "error_message":     "Error Message",
}

_SUPPORTED_EXTENSIONS: frozenset[str] = frozenset({".xlsx", ".xls"})

# ── Colour palette ─────────────────────────────────────────────────────────────
_COL_HEADER_BG      = "1F3864"   # Dark navy — header background
_COL_HEADER_FG      = "FFFFFF"   # White — header text
_COL_SUCCESS_BG     = "E2EFDA"   # Light green — SUCCESS status
_COL_SUCCESS_FG     = "375623"   # Dark green — SUCCESS text
_COL_FAILED_BG      = "FCE4D6"   # Light red — FAILED status
_COL_FAILED_FG      = "833C00"   # Dark red — FAILED text
_COL_ERROR_COUNT_BG = "FFF2CC"   # Amber — non-zero error count
_COL_ERROR_COUNT_FG = "7F6000"   # Dark amber — non-zero error count text
_COL_ALT_ROW_BG     = "F2F7FF"   # Very light blue — alternate rows
_COL_BORDER         = "BDD7EE"   # Light blue — cell borders

# ── Column width overrides (in characters) ────────────────────────────────────
_FIXED_WIDTHS: dict[str, int] = {
    "query":             60,
    "total_count_query": 55,
    "error_count_query": 55,
    "total_count":       14,
    "error_count":       14,
    "status":            12,
    "error_message":     40,
}

_SQL_COLUMNS = {"query", "total_count_query", "error_count_query"}


# ── Public API ─────────────────────────────────────────────────────────────────

def read_queries(filepath: str) -> list[str]:
    """
    Read SQL queries from the input Excel file.

    Expects a column named ``query`` (case-sensitive).
    Blank rows and whitespace-only cells are silently skipped.

    Args:
        filepath: Path to the input ``.xlsx`` / ``.xls`` file.

    Returns:
        List of non-empty SQL strings.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file format is unsupported or the required
                    column is missing.
    """
    path = Path(filepath)

    if not path.exists():
        raise FileNotFoundError(
            f"Input Excel file not found: '{filepath}'. "
            "Please place your queries file at this location."
        )

    if path.suffix.lower() not in _SUPPORTED_EXTENSIONS:
        raise ValueError(
            f"Unsupported file extension '{path.suffix}'. "
            f"Supported formats: {sorted(_SUPPORTED_EXTENSIONS)}"
        )

    logger.info("Reading input Excel file: %s", filepath)

    try:
        df = pd.read_excel(filepath, engine="openpyxl", dtype=str)
    except Exception as exc:
        raise ValueError(
            f"Failed to read Excel file '{filepath}': {exc}"
        ) from exc

    if INPUT_QUERY_COLUMN not in df.columns:
        raise ValueError(
            f"Required column '{INPUT_QUERY_COLUMN}' not found in '{filepath}'. "
            f"Columns present: {list(df.columns)}"
        )

    # Filter: drop NaN and whitespace-only entries
    raw_series = df[INPUT_QUERY_COLUMN].dropna().astype(str).str.strip()
    queries = [q for q in raw_series if q]

    logger.info("Loaded %d non-empty queries from '%s'.", len(queries), filepath)
    return queries


def write_results(filepath: str, results: list[dict[str, Any]]) -> None:
    """
    Write processing results to a professionally styled output Excel file.

    Styling includes:
    - Dark navy header row with white bold text
    - Alternating row shading
    - Green / red status cell colouring
    - Amber highlight for non-zero error counts
    - Wrapped SQL text in query columns
    - Frozen header row + first column
    - Thin borders on all cells
    - Auto-sized column widths

    Args:
        filepath: Destination ``.xlsx`` path.
        results:  List of result dicts (keys must include OUTPUT_COLUMNS).

    Raises:
        ValueError: If the file cannot be written.
    """
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)

    df = pd.DataFrame(results, columns=OUTPUT_COLUMNS)

    logger.info("Writing %d result rows to '%s'.", len(df), filepath)

    try:
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Write data (use friendly column names as headers)
            df_display = df.copy()
            df_display.columns = [COLUMN_HEADERS[c] for c in OUTPUT_COLUMNS]
            df_display.to_excel(writer, index=False, sheet_name="Validation Results")

            ws = writer.sheets["Validation Results"]
            _apply_styles(ws, df)

    except Exception as exc:
        raise ValueError(
            f"Failed to write output Excel file '{filepath}': {exc}"
        ) from exc

    logger.info("Results successfully written to '%s'.", filepath)


# ── Internal styling helpers ───────────────────────────────────────────────────

def _thin_border() -> Border:
    side = Side(style="thin", color=_COL_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _apply_styles(ws: Any, df: pd.DataFrame) -> None:
    """Apply all visual styling to the worksheet."""
    n_rows = len(df)
    n_cols = len(OUTPUT_COLUMNS)
    border  = _thin_border()

    # ── 1. Header row styling ─────────────────────────────────────────────
    header_fill = _fill(_COL_HEADER_BG)
    header_font = Font(
        name="Calibri", bold=True, color=_COL_HEADER_FG, size=11
    )
    header_align = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.row_dimensions[1].height = 32

    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = header_font and header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = border

    # ── 2. Data rows ──────────────────────────────────────────────────────
    col_names = OUTPUT_COLUMNS  # original column names

    for row_idx in range(2, n_rows + 2):
        data_row_idx = row_idx - 2           # 0-based index into df
        is_alt = data_row_idx % 2 == 1
        row_bg = _fill(_COL_ALT_ROW_BG) if is_alt else _fill("FFFFFF")

        status_val = str(df.iloc[data_row_idx].get("status", ""))
        error_val  = df.iloc[data_row_idx].get("error_count", None)

        # Row height — taller for SQL rows
        ws.row_dimensions[row_idx].height = 70

        for col_idx, col_name in enumerate(col_names, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border

            # ── Per-column formatting ─────────────────────────────────
            if col_name in _SQL_COLUMNS:
                # SQL text: wrap, top-aligned, monospace-style
                cell.alignment = Alignment(
                    wrap_text=True, vertical="top", horizontal="left"
                )
                cell.font = Font(name="Courier New", size=9)
                cell.fill = row_bg

            elif col_name == "status":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Calibri", bold=True, size=10)
                if status_val == "SUCCESS":
                    cell.fill = _fill(_COL_SUCCESS_BG)
                    cell.font = Font(
                        name="Calibri", bold=True, size=10,
                        color=_COL_SUCCESS_FG
                    )
                elif status_val == "FAILED":
                    cell.fill = _fill(_COL_FAILED_BG)
                    cell.font = Font(
                        name="Calibri", bold=True, size=10,
                        color=_COL_FAILED_FG
                    )
                else:
                    cell.fill = row_bg

            elif col_name == "error_count":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Calibri", size=11, bold=True)
                try:
                    numeric_val = int(float(str(error_val))) if error_val not in (None, "None", "") else None
                except (ValueError, TypeError):
                    numeric_val = None

                if numeric_val is not None and numeric_val > 0:
                    # Highlight non-zero error counts in amber
                    cell.fill = _fill(_COL_ERROR_COUNT_BG)
                    cell.font = Font(
                        name="Calibri", size=11, bold=True,
                        color=_COL_ERROR_COUNT_FG
                    )
                else:
                    cell.fill = row_bg

            elif col_name == "total_count":
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name="Calibri", size=11)
                cell.fill = row_bg

            elif col_name == "error_message":
                cell.alignment = Alignment(
                    wrap_text=True, vertical="top", horizontal="left"
                )
                cell.font = Font(name="Calibri", size=10, italic=True, color="C00000")
                cell.fill = row_bg

            else:
                cell.alignment = Alignment(
                    wrap_text=True, vertical="top", horizontal="left"
                )
                cell.font = Font(name="Calibri", size=10)
                cell.fill = row_bg

    # ── 3. Column widths ──────────────────────────────────────────────────
    for col_idx, col_name in enumerate(col_names, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _FIXED_WIDTHS.get(col_name, 20)

    # ── 4. Freeze header row ──────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── 5. Sheet tab colour ───────────────────────────────────────────────
    ws.sheet_properties.tabColor = "1F3864"

    # ── 6. Add a summary row at the bottom ────────────────────────────────
    summary_row = n_rows + 3
    total_col   = col_names.index("total_count") + 1
    error_col   = col_names.index("error_count") + 1
    status_col  = col_names.index("status") + 1

    # Label
    label_cell = ws.cell(row=summary_row, column=1, value="SUMMARY")
    label_cell.font = Font(name="Calibri", bold=True, size=11, color=_COL_HEADER_FG)
    label_cell.fill = _fill(_COL_HEADER_BG)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = _thin_border()

    successful = sum(1 for r in range(len(df)) if str(df.iloc[r]["status"]) == "SUCCESS")
    failed     = len(df) - successful
    queries_with_errors = sum(
        1 for r in range(len(df))
        if df.iloc[r]["error_count"] not in (None, "None", "")
        and str(df.iloc[r]["error_count"]) != "0"
    )

    summary_text = (
        f"Total Queries: {len(df)}   |   "
        f"SUCCESS: {successful}   |   "
        f"FAILED: {failed}   |   "
        f"Queries with Data Issues: {queries_with_errors}"
    )
    summary_cell = ws.cell(row=summary_row, column=2, value=summary_text)
    summary_cell.font = Font(name="Calibri", bold=True, size=11, color=_COL_HEADER_FG)
    summary_cell.fill = _fill(_COL_HEADER_BG)
    summary_cell.alignment = Alignment(horizontal="left", vertical="center")
    summary_cell.border = _thin_border()

    # Merge across remaining columns
    ws.merge_cells(
        start_row=summary_row, start_column=2,
        end_row=summary_row,   end_column=n_cols
    )
    ws.row_dimensions[summary_row].height = 24
